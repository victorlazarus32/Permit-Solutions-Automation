"""
Export a clean, review-ready spreadsheet of every filtered violation in the DB.

Splits into three sheets so Victor can decide what to mail:
  1. "Ready to Mail"      -- has owner + mailing address + not yet mailed
  2. "Needs Owner Lookup" -- missing owner or address (held back from mailing)
  3. "Already Mailed"     -- lob_letter_id is set (proof of work)

Run:
    python -m scripts.export_review

Output:
    data/leads_for_review_<YYYY-MM-DD-HHMM>.xlsx
"""
from __future__ import annotations

import sqlite3
import sys
from datetime import datetime, date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from db import DB_PATH

# Map source code -> friendly display name
SOURCE_PRETTY = {
    "miami_dade_unincorporated": "Miami-Dade Unincorporated",
    "homestead": "Homestead",
}

# Columns we surface in every sheet, in this order
DISPLAY_COLS = [
    "Source",
    "Case Number",
    "Open Date",
    "Days Open",
    "Property Address",
    "Owner",
    "Owner Mailing Address",
    "Matched Keywords",
    "Violation Summary",
    "Notes",
]


def fetch_violations() -> pd.DataFrame:
    """Pull every keyword-matched violation from the DB."""
    with sqlite3.connect(DB_PATH) as conn:
        df = pd.read_sql_query(
            """
            SELECT source, case_number, open_date, property_address,
                   owner_full_name, owner_mailing_address,
                   matched_keywords, alleged_violation, comments,
                   lob_letter_id, lob_status, lob_mailed_at
            FROM violations
            ORDER BY open_date DESC, case_number
            """,
            conn,
        )
    return df


def project(df: pd.DataFrame) -> pd.DataFrame:
    """Reshape raw DB rows into the display columns."""
    today = date.today()

    def _days_open(d):
        if not d:
            return None
        try:
            return (today - date.fromisoformat(d)).days
        except (ValueError, TypeError):
            return None

    def _summary(av):
        if not av or pd.isna(av):
            return ""
        s = str(av).replace("\n", " ").strip()
        return s if len(s) <= 250 else s[:247] + "..."

    out = pd.DataFrame({
        "Source":                df["source"].map(lambda s: SOURCE_PRETTY.get(s, s)),
        "Case Number":           df["case_number"],
        "Open Date":             pd.to_datetime(df["open_date"], errors="coerce").dt.date,
        "Days Open":             df["open_date"].map(_days_open),
        "Property Address":      df["property_address"].fillna(""),
        "Owner":                 df["owner_full_name"].fillna(""),
        "Owner Mailing Address": df["owner_mailing_address"].fillna(""),
        "Matched Keywords":      df["matched_keywords"].fillna(""),
        "Violation Summary":     df["alleged_violation"].map(_summary),
        "Notes":                 df["comments"].fillna(""),
    })
    return out


def split_buckets(df_raw: pd.DataFrame, df_display: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Return {sheet_name: dataframe} for each of the 3 review buckets."""
    has_owner = df_raw["owner_full_name"].notna() & (df_raw["owner_full_name"].astype(str).str.strip() != "")
    has_addr = df_raw["owner_mailing_address"].notna() & (df_raw["owner_mailing_address"].astype(str).str.strip() != "")
    flagged = df_raw["comments"].fillna("").str.contains("NEEDS_OWNER_LOOKUP", case=False)
    mailed = df_raw["lob_letter_id"].notna()

    ready_mask = has_owner & has_addr & ~flagged & ~mailed
    needs_mask = (~has_owner | ~has_addr | flagged) & ~mailed
    mailed_mask = mailed

    return {
        "Ready to Mail":      df_display[ready_mask].reset_index(drop=True),
        "Needs Owner Lookup": df_display[needs_mask].reset_index(drop=True),
        "Already Mailed":     df_display[mailed_mask].reset_index(drop=True),
    }


def write_workbook(buckets: dict[str, pd.DataFrame], out_path: Path) -> None:
    """Write all sheets with formatting (bold header, frozen top row, auto-width)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in buckets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            _format_sheet(ws, df)


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Bold header, freeze pane, sensible column widths."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3A52", end_color="1F3A52", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"

    # Width per column based on content (capped so the violation column does not blow up)
    width_caps = {
        "Source": 26,
        "Case Number": 14,
        "Open Date": 12,
        "Days Open": 10,
        "Property Address": 32,
        "Owner": 30,
        "Owner Mailing Address": 38,
        "Matched Keywords": 28,
        "Violation Summary": 80,
        "Notes": 30,
    }
    for idx, col in enumerate(df.columns, start=1):
        cap = width_caps.get(col, 24)
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].head(200).tolist()])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, cap)

    # Wrap long text in violation + mailing address columns
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in {
                get_column_letter(list(df.columns).index("Violation Summary") + 1) if "Violation Summary" in df.columns else "",
                get_column_letter(list(df.columns).index("Owner Mailing Address") + 1) if "Owner Mailing Address" in df.columns else "",
                get_column_letter(list(df.columns).index("Notes") + 1) if "Notes" in df.columns else "",
            }:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def main() -> Path:
    df_raw = fetch_violations()
    df_display = project(df_raw)
    buckets = split_buckets(df_raw, df_display)

    stamp = datetime.now().strftime("%Y-%m-%d-%H%M")
    out_path = PROJECT_ROOT / "data" / f"leads_for_review_{stamp}.xlsx"
    write_workbook(buckets, out_path)

    print(f"Wrote {out_path}")
    for name, df in buckets.items():
        print(f"  {name:22s} {len(df):4d} rows")
    return out_path


if __name__ == "__main__":
    main()
